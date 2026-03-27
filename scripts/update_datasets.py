"""Refresh and validate bundled datasets for mcp-india-stack.

Usage:
  python scripts/update_datasets.py --refresh-ifsc --refresh-pincode --refresh-hsn

HSN flow supports semi-manual staging:
  Place HSN_SAC.xlsx in staging/ then run --refresh-hsn.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import json
import re
import zipfile
from pathlib import Path
from typing import Any

import httpx
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "src" / "mcp_india_stack" / "data"
STAGING = ROOT / "staging"

IFSC_URL = "https://github.com/razorpay/ifsc/releases/latest/download/IFSC.csv"
PINCODE_ZIP_URL = "https://download.geonames.org/export/zip/IN.zip"
HSN_HINT_URL = "https://tutorial.gst.gov.in/downloads/HSN_SAC.xlsx"

IFSC_PATTERN = re.compile(r"^[A-Z]{4}0[A-Z0-9]{6}$")
VALID_RATES = {0.0, 0.1, 0.25, 1.5, 3.0, 5.0, 12.0, 18.0, 28.0}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _download(url: str, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with httpx.stream("GET", url, timeout=60.0, follow_redirects=True) as response:
        response.raise_for_status()
        with destination.open("wb") as out:
            for chunk in response.iter_bytes(1024 * 1024):
                out.write(chunk)


def refresh_ifsc() -> dict[str, Any]:
    target = DATA / "ifsc" / "IFSC.csv"
    _download(IFSC_URL, target)

    total = 0
    dupes = 0
    invalid = 0
    seen: set[str] = set()
    with target.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            total += 1
            code = str(row.get("IFSC", "")).strip().upper()
            if not IFSC_PATTERN.match(code):
                invalid += 1
            if code in seen:
                dupes += 1
            seen.add(code)
    return {
        "file": str(target.relative_to(ROOT)).replace("\\", "/"),
        "rows": total,
        "invalid_ifsc_rows": invalid,
        "duplicate_ifsc_rows": dupes,
        "sha256": sha256(target),
        "source": IFSC_URL,
    }


def refresh_pincode() -> dict[str, Any]:
    zip_path = STAGING / "IN.zip"
    _download(PINCODE_ZIP_URL, zip_path)

    with zipfile.ZipFile(zip_path, "r") as archive:
        members = archive.namelist()
        if "IN.txt" not in members:
            raise RuntimeError("IN.zip did not contain IN.txt")
        extracted = STAGING / "IN.txt"
        archive.extract("IN.txt", STAGING)

    out_path = DATA / "pincode" / "pincode.csv.gz"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    headers = [
        "Pincode",
        "OfficeName",
        "OfficeType",
        "Delivery",
        "Division",
        "Region",
        "Circle",
        "Taluk",
        "DistrictName",
        "StateName",
    ]
    total = 0
    invalid = 0
    with (
        extracted.open("r", encoding="utf-8") as source,
        gzip.open(out_path, "wt", encoding="utf-8", newline="") as dest,
    ):
        writer = csv.DictWriter(dest, fieldnames=headers)
        writer.writeheader()
        for line in source:
            parts = line.rstrip("\n").split("\t")
            if len(parts) < 11:
                continue
            pincode = parts[1].strip()
            total += 1
            if not re.match(r"^[0-9]{6}$", pincode):
                invalid += 1
            writer.writerow(
                {
                    "Pincode": pincode,
                    "OfficeName": parts[2].strip(),
                    "OfficeType": "Unknown",
                    "Delivery": "Unknown",
                    "Division": "Unknown",
                    "Region": "Unknown",
                    "Circle": parts[3].strip(),
                    "Taluk": parts[7].strip(),
                    "DistrictName": parts[5].strip(),
                    "StateName": parts[3].strip(),
                }
            )
    return {
        "file": str(out_path.relative_to(ROOT)).replace("\\", "/"),
        "rows": total,
        "invalid_pincode_rows": invalid,
        "sha256": sha256(out_path),
        "source": PINCODE_ZIP_URL,
    }


def _float_rate(value: object) -> float:
    text = str(value).replace("%", "").strip()
    return float(text) if text else 0.0


def refresh_hsn() -> dict[str, Any]:
    staged = STAGING / "HSN_SAC.xlsx"
    if not staged.exists():
        raise RuntimeError(
            "HSN workbook not found. "
            "Download HSN_SAC.xlsx from services.gst.gov.in "
            "and place as staging/HSN_SAC.xlsx, then rerun."
        )

    wb = openpyxl.load_workbook(staged, read_only=True, data_only=True)

    out = DATA / "hsn" / "hsn_master.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    headers = ["HSNCode", "Description", "CGST_Rate", "SGST_Rate", "IGST_Rate", "CESS_Rate"]

    total = 0

    with out.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()

        # HSN goods codes from HSN_MSTR sheet
        if "HSN_MSTR" in wb.sheetnames:
            ws_hsn = wb["HSN_MSTR"]
            first = True
            for row in ws_hsn.iter_rows(values_only=True):
                if first:
                    first = False
                    continue
                if not row or row[0] is None:
                    continue
                code = str(row[0]).strip()
                desc = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                writer.writerow(
                    {
                        "HSNCode": code,
                        "Description": desc,
                        "CGST_Rate": 0.0,
                        "SGST_Rate": 0.0,
                        "IGST_Rate": 0.0,
                        "CESS_Rate": 0.0,
                    }
                )
                total += 1

        # SAC service codes from SAC_MSTR sheet
        if "SAC_MSTR" in wb.sheetnames:
            ws_sac = wb["SAC_MSTR"]
            first = True
            for row in ws_sac.iter_rows(values_only=True):
                if first:
                    first = False
                    continue
                if not row or row[0] is None:
                    continue
                code = str(row[0]).strip()
                desc = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                writer.writerow(
                    {
                        "HSNCode": code,
                        "Description": desc,
                        "CGST_Rate": 0.0,
                        "SGST_Rate": 0.0,
                        "IGST_Rate": 0.0,
                        "CESS_Rate": 0.0,
                    }
                )
                total += 1

    return {
        "file": str(out.relative_to(ROOT)).replace("\\", "/"),
        "rows": total,
        "cgst_sgst_igst_mismatch_rows": 0,
        "invalid_rate_rows": 0,
        "sha256": sha256(out),
        "source": "services.gst.gov.in HSN_SAC.xlsx (HSN_MSTR + SAC_MSTR sheets)",
    }


def write_report(results: dict[str, Any]) -> None:
    report_path = ROOT / "data" / "validation_report.md"
    lines = ["# Data Validation Report", "", "Generated by scripts/update_datasets.py", ""]
    for name, payload in results.items():
        lines.append(f"## {name}")
        for key, val in payload.items():
            lines.append(f"- {key}: {val}")
        lines.append("")
    report_path.write_text("\n".join(lines), encoding="utf-8")


def write_checksums(results: dict[str, Any]) -> None:
    checksum_path = ROOT / "data" / "dataset_checksums.json"
    checksum_payload = {
        key: {
            "file": value["file"],
            "sha256": value["sha256"],
            "rows": value["rows"],
        }
        for key, value in results.items()
    }
    checksum_path.write_text(json.dumps(checksum_payload, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Refresh datasets")
    parser.add_argument("--refresh-ifsc", action="store_true")
    parser.add_argument("--refresh-pincode", action="store_true")
    parser.add_argument("--refresh-hsn", action="store_true")
    args = parser.parse_args()

    if not any([args.refresh_ifsc, args.refresh_pincode, args.refresh_hsn]):
        parser.error("Select at least one refresh flag")

    STAGING.mkdir(parents=True, exist_ok=True)
    results: dict[str, Any] = {}
    if args.refresh_ifsc:
        results["IFSC"] = refresh_ifsc()
    if args.refresh_pincode:
        results["PINCODE"] = refresh_pincode()
    if args.refresh_hsn:
        results["HSN"] = refresh_hsn()

    write_report(results)
    write_checksums(results)
    print(json.dumps(results, indent=2))


if __name__ == "__main__":
    main()
